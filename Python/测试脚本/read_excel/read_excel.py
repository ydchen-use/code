import openpyxl
import pandas as pd


def read_excel(file_path):
    """
    文件
    :param file_path: 文件路径
    :return:
    """
    resault = ""
    wb = openpyxl.load_workbook(file_path)
    # 获取文件中的所有表格
    sheets = wb.sheetnames
    for i in range(len(sheets)):
        sheet = wb[sheets[i]]
        for r in range(1, sheet.max_row + 1):
            if r == 1:
                resault = resault + "\n" + "".join(
                    [str(sheet.cell(row=r, column=c).value).ljust(17) for c in range(1, sheet.max_column + 1)])
            else:
                resault = resault + "".join(
                    [str(sheet.cell(row=r, column=c).value).ljust(20) for c in range(1, sheet.max_column + 1)])

    return resault


def read_excel_pandas(file_path):
    """
    用 pandas 读取 excel 文件
    :param file_path:
    :return:
    """
    data = pd.read_excel(file_path)
    rules_sids_dict = dict(zip(data["规则Id"], data["事件分类"]))
    return rules_sids_dict


def get_malware_sid(file_path):
    """

    :param file_path:
    :return:
    """
    malware_rules_sid_list = []
    data = pd.read_excel(file_path)

    rules_sids_dict = dict(zip(data["规则Id"], data["事件分类"]))

    for key in rules_sids_dict:
        if rules_sids_dict[key] == "恶意软件":
            malware_rules_sid_list.append(key)

    return malware_rules_sid_list


if __name__ == "__main__":
    read_excel_pandas("rules_manager.xlsm")


